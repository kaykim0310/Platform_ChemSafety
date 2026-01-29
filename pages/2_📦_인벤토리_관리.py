"""
화학물질 인벤토리 관리 시스템
- KOSHA API 연동 (8번: 노출기준, 15번: 법적규제+위험물)
- KECO API 연동 (환경부 규제)
- PRTR 배출량조사 대상물질 체크
"""
import streamlit as st
import pandas as pd
import io
from pathlib import Path

# 경로 설정
current_dir = Path(__file__).parent.parent
import sys
sys.path.insert(0, str(current_dir))

# KOSHA API
try:
    from core.kosha_api import get_chemical_info
    KOSHA_AVAILABLE = True
except ImportError:
    KOSHA_AVAILABLE = False

# KECO API
try:
    from core.keco_api import get_chemical_regulations
    KECO_AVAILABLE = True
except ImportError:
    KECO_AVAILABLE = False

# PRTR DB
try:
    from core.prtr_db import check_prtr_status
except ImportError:
    def check_prtr_status(cas):
        return {"대상여부": "-", "그룹": "-", "기준취급량": "-"}

st.set_page_config(page_title="인벤토리 관리", page_icon="📦", layout="wide")

# CSS
st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1.5rem;
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        border-radius: 10px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .upload-box {
        border: 2px dashed #94a3b8;
        border-radius: 10px;
        padding: 2rem;
        text-align: center;
        background: #f8fafc;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# 세션 상태
if 'inventory' not in st.session_state:
    st.session_state.inventory = []

# ============================================
# 함수들
# ============================================
def query_chemical_info(cas_no):
    """KOSHA API 조회"""
    if not KOSHA_AVAILABLE:
        return None, "KOSHA API 없음"
    try:
        result = get_chemical_info(cas_no)
        if result.get('success'):
            return result, None
        else:
            return None, result.get('error', '조회 실패')
    except Exception as e:
        return None, f"API 오류: {str(e)[:50]}"

def create_inventory_item(process_name, unit_workplace, product_name, chem_name, alias, cas_no, content, kosha_data=None, keco_data=None, prtr_status=None):
    """인벤토리 항목 생성"""
    item = {
        '공정명': process_name or '',
        '단위작업장소': unit_workplace or '',
        '제품명': product_name or '',
        '화학물질명': chem_name or '',
        '관용명/이명': alias or '',
        'CAS No': cas_no or '',
        '함유량(%)': content or '',
        # 독성정보 (KOSHA 8번)
        '발암성': '-', '변이성': '-', '생식독성': '-', '노출기준(TWA)': '-',
        # 산안법 규제 (KOSHA 15번)
        '작업환경측정': 'X', '특수건강진단': 'X', '관리대상유해물질': 'X', '특별관리물질': 'X',
        # 위험물 (KOSHA 15번)
        '위험물류별': '-', '지정수량': '-', '위험등급': '-',
        # 환경부 규제 (KECO API)
        '기존': '-', '급성·만성·생태': 'X', '사고대비': 'X', '제한/금지/허가': '-',
        '중점': '-', '잔류': '-', '함량 및 규제정보': '-', '등록대상기존화학물질': '-', '기존물질여부': '-',
        # PRTR
        'PRTR그룹': '-', 'PRTR기준량': '-'
    }
    
    # KOSHA 데이터 적용
    if kosha_data:
        if not chem_name:
            item['화학물질명'] = kosha_data.get('chemNmKr', '') or kosha_data.get('chemNmEn', '')
        # 8번 항목: 노출기준
        item['노출기준(TWA)'] = kosha_data.get('twa', '-') or '-'
        
        # 15번 항목: 산안법 규제
        item['작업환경측정'] = kosha_data.get('workMeasure', 'X') or 'X'
        item['특수건강진단'] = kosha_data.get('specialHealth', 'X') or 'X'
        item['관리대상유해물질'] = kosha_data.get('managedSubstance', 'X') or 'X'
        item['특별관리물질'] = kosha_data.get('specialManaged', 'X') or 'X'
        
        # 15번 항목: 위험물안전관리법
        hazmat_class = kosha_data.get('hazmatClass', '-') or '-'
        hazmat_name = kosha_data.get('hazmatName', '-') or '-'
        if hazmat_class != '-' and hazmat_name != '-':
            item['위험물류별'] = f"{hazmat_class} {hazmat_name}"
        elif hazmat_class != '-':
            item['위험물류별'] = hazmat_class
        item['지정수량'] = kosha_data.get('hazmatQuantity', '-') or '-'
        item['위험등급'] = kosha_data.get('hazmatGrade', '-') or '-'
    
    # KECO 데이터 적용 (환경부)
    if keco_data and keco_data.get('success'):
        data = keco_data
        
        existing = data.get('기존화학물질', '-')
        if existing and existing != '-':
            item['기존'] = 'O'
            item['기존물질여부'] = existing
        
        toxic = data.get('유독물질', '-')
        human_hazard = data.get('인체유해성물질', '-')
        if toxic and toxic != '-':
            item['급성·만성·생태'] = toxic
        elif human_hazard and human_hazard != '-':
            item['급성·만성·생태'] = human_hazard
        
        accident = data.get('사고대비물질', '-')
        if accident and accident != '-':
            item['사고대비'] = accident
        
        reg_list = []
        restricted = data.get('제한물질', '-')
        prohibited = data.get('금지물질', '-')
        permitted = data.get('허가물질', '-')
        if restricted and restricted != '-':
            reg_list.append(f"제한({restricted})")
        if prohibited and prohibited != '-':
            reg_list.append(f"금지({prohibited})")
        if permitted and permitted != '-':
            reg_list.append(f"허가({permitted})")
        if reg_list:
            item['제한/금지/허가'] = ', '.join(reg_list)
        
        priority = data.get('중점관리물질', '-')
        if priority and priority != '-':
            item['중점'] = priority
        
        reg_existing = data.get('등록대상기존화학물질', '-')
        if reg_existing and reg_existing != '-':
            item['등록대상기존화학물질'] = reg_existing
        
        details = data.get('details', {})
        if details:
            info_list = []
            for k, v in details.items():
                if '함량' in k:
                    info_list.append(f"{k}: {v}")
            if info_list:
                item['함량 및 규제정보'] = '; '.join(info_list[:2])
    
    # PRTR
    if prtr_status and prtr_status.get('대상여부') == 'O':
        item['PRTR그룹'] = prtr_status.get('그룹', '-')
        item['PRTR기준량'] = prtr_status.get('기준취급량', '-')
    
    return item

def create_template_excel():
    """빈 템플릿 생성"""
    output = io.BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "화학물질 정보"
    
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    header_fill2 = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'] = '공정명', '단위작업장소', '제품명', '화학물질명', '관용명/이명', 'CAS No', '함유량(%)'
    ws['H1'], ws['L1'], ws['P1'], ws['S1'] = '독성정보', '법적규제 대상여부', '위험물', '환경부 법적규제 대상여부'
    
    row2 = ['', '', '', '', '', '', '', '발암성', '변이성', '생식독성', '노출기준(TWA)', '작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질', '위험물류별', '지정수량', '위험등급', '기존', '급성·만성·생태', '사고대비', '제한/금지/허가', '중점', '잔류', '함량 및 규제정보', '등록대상기존화학물질', '기존물질여부']
    for col, h in enumerate(row2, 1):
        ws.cell(row=2, column=col, value=h)
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.merge_cells(f'{col}1:{col}2')
    ws.merge_cells('H1:K1')
    ws.merge_cells('L1:O1')
    ws.merge_cells('P1:R1')
    ws.merge_cells('S1:AA1')
    
    for row in [1, 2]:
        for col in range(1, 28):
            cell = ws.cell(row=row, column=col)
            cell.font, cell.alignment, cell.border = header_font, center_align, thin_border
            cell.fill = header_fill if row == 1 else header_fill2
    
    wb.save(output)
    output.seek(0)
    return output

def export_inventory_to_excel(inventory_data):
    """인벤토리 내보내기"""
    output = io.BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "화학물질 정보"
    
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    header_fill2 = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    yes_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    hazmat_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'] = '공정명', '단위작업장소', '제품명', '화학물질명', '관용명/이명', 'CAS No', '함유량(%)'
    ws['H1'], ws['L1'], ws['P1'], ws['S1'] = '독성정보', '법적규제 대상여부', '위험물', '환경부 법적규제 대상여부'
    
    row2 = ['', '', '', '', '', '', '', '발암성', '변이성', '생식독성', '노출기준(TWA)', '작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질', '위험물류별', '지정수량', '위험등급', '기존', '급성·만성·생태', '사고대비', '제한/금지/허가', '중점', '잔류', '함량 및 규제정보', '등록대상기존화학물질', '기존물질여부']
    for col, h in enumerate(row2, 1):
        ws.cell(row=2, column=col, value=h)
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.merge_cells(f'{col}1:{col}2')
    ws.merge_cells('H1:K1')
    ws.merge_cells('L1:O1')
    ws.merge_cells('P1:R1')
    ws.merge_cells('S1:AA1')
    
    for row in [1, 2]:
        for col in range(1, 28):
            cell = ws.cell(row=row, column=col)
            cell.font, cell.alignment, cell.border = header_font, center_align, thin_border
            cell.fill = header_fill if row == 1 else header_fill2
    
    for row_idx, item in enumerate(inventory_data, 3):
        data = [
            item.get('공정명', ''), item.get('단위작업장소', ''), item.get('제품명', ''),
            item.get('화학물질명', ''), item.get('관용명/이명', ''), item.get('CAS No', ''), item.get('함유량(%)', ''),
            item.get('발암성', '-'), item.get('변이성', '-'), item.get('생식독성', '-'), item.get('노출기준(TWA)', '-'),
            item.get('작업환경측정', 'X'), item.get('특수건강진단', 'X'), item.get('관리대상유해물질', 'X'), item.get('특별관리물질', 'X'),
            item.get('위험물류별', '-'), item.get('지정수량', '-'), item.get('위험등급', '-'),
            item.get('기존', '-'), item.get('급성·만성·생태', 'X'), item.get('사고대비', 'X'), item.get('제한/금지/허가', '-'),
            item.get('중점', '-'), item.get('잔류', '-'), item.get('함량 및 규제정보', '-'), item.get('등록대상기존화학물질', '-'), item.get('기존물질여부', '-')
        ]
        
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            if val == 'O':
                cell.fill = yes_fill
            if col_idx in [16, 17, 18] and val not in ['-', '', None]:
                cell.fill = hazmat_fill
    
    col_widths = {'A': 10, 'B': 12, 'C': 18, 'D': 18, 'E': 12, 'F': 12, 'G': 10, 'H': 10, 'I': 8, 'J': 8, 'K': 12, 'L': 10, 'M': 10, 'N': 12, 'O': 10, 'P': 25, 'Q': 10, 'R': 8, 'S': 6, 'T': 6, 'U': 8, 'V': 12, 'W': 6, 'X': 6, 'Y': 12, 'Z': 14}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output)
    output.seek(0)
    return output

# ============================================
# 사이드바
# ============================================
with st.sidebar:
    st.header("📦 인벤토리 관리")
    st.metric("등록된 물질", f"{len(st.session_state.inventory)}종")
    
    if len(st.session_state.inventory) > 0:
        cnt = sum(1 for i in st.session_state.inventory if i.get('작업환경측정') == 'O')
        st.metric("측정대상", f"{cnt}종")
        cnt_haz = sum(1 for i in st.session_state.inventory if i.get('위험물류별', '-') != '-')
        st.metric("위험물", f"{cnt_haz}종")
    
    st.divider()
    st.subheader("📥 템플릿")
    template = create_template_excel()
    st.download_button("📄 템플릿 다운로드", template, "template_inventory.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    st.divider()
    st.caption(f"KOSHA API: {'✅' if KOSHA_AVAILABLE else '❌'} (고용노동부)")
    st.caption(f"KECO API: {'✅' if KECO_AVAILABLE else '❌'} (환경부)")
    st.caption("PRTR DB: ✅ (배출량조사)")
    
    st.divider()
    if st.button("🗑️ 전체 삭제", use_container_width=True):
        st.session_state.inventory = []
        st.rerun()

# ============================================
# 메인
# ============================================
st.markdown("""
<div class="main-header">
    <h2>📦 화학물질 인벤토리 관리</h2>
    <p>KOSHA API + KECO API + PRTR DB 연동</p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs(["📤 엑셀 업로드", "➕ 개별 등록", "📋 목록 보기", "📥 내보내기"])

# ============================================
# 탭 1: 엑셀 업로드
# ============================================
with tab1:
    st.subheader("📤 엑셀 파일 업로드")
    
    st.markdown('<div class="upload-box"><h4>📁 엑셀 파일을 업로드하세요</h4></div>', unsafe_allow_html=True)
    st.warning("⚠️ **API 호출 제한**: 한 번에 최대 **300건**까지 처리됩니다. 대용량 파일은 여러 번 나눠서 등록해주세요.")
    
    uploaded_file = st.file_uploader("엑셀 파일 선택", type=['xlsx', 'xls'])
    
    if uploaded_file:
        st.success(f"✅ **{uploaded_file.name}** 업로드됨")
        
        try:
            # openpyxl iter_rows로 빠르게 전체 읽기
            from openpyxl import load_workbook
            
            wb = load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
            ws = wb.active
            
            # 헤더 읽기 - 1행과 2행 모두 확인 (병합셀 처리)
            headers = []
            row1_cells = list(ws[1])
            row2_cells = list(ws[2])
            
            for i in range(len(row1_cells)):
                val1 = row1_cells[i].value
                val2 = row2_cells[i].value if i < len(row2_cells) else None
                
                # 1행 값이 있으면 1행 사용, 없으면 2행 사용
                if val1 and str(val1).strip():
                    headers.append(str(val1).strip())
                elif val2 and str(val2).strip():
                    headers.append(str(val2).strip())
                else:
                    headers.append(f"Col_{i+1}")
            
            # 데이터 읽기 (3행부터)
            data_rows = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and any(cell is not None for cell in row):
                    row_dict = {}
                    for i in range(min(len(headers), len(row))):
                        row_dict[headers[i]] = row[i]
                    data_rows.append(row_dict)
            
            wb.close()
            
            df = pd.DataFrame(data_rows)
            df = df.dropna(how='all')
            
            with st.expander("📊 미리보기", expanded=True):
                st.dataframe(df.head(10), use_container_width=True)
                st.caption(f"총 **{len(df)}행**")
            
            st.divider()
            
            # 컬럼 자동 매칭 함수
            def find_column(df_columns, keywords):
                """키워드로 컬럼 찾기"""
                for col in df_columns:
                    col_lower = str(col).lower()
                    for kw in keywords:
                        if kw in col_lower:
                            return col
                return None
            
            # 자동 매칭
            auto_cas = find_column(df.columns, ['cas'])
            auto_name = find_column(df.columns, ['화학물질명', '물질명', 'chemical'])
            auto_process = find_column(df.columns, ['공정명', '공정'])
            auto_unit = find_column(df.columns, ['단위작업장소', '단위작업', '작업장소'])
            auto_product = find_column(df.columns, ['제품명', '제품'])
            auto_content = find_column(df.columns, ['함유량', '함량', '농도', 'content'])
            
            col1, col2 = st.columns(2)
            with col1:
                cas_col = st.selectbox("CAS 번호 컬럼 *", list(df.columns), index=list(df.columns).index(auto_cas) if auto_cas else 0)
                name_col = st.selectbox("화학물질명 컬럼", ['(자동조회)'] + list(df.columns), index=(['(자동조회)'] + list(df.columns)).index(auto_name) if auto_name else 0)
            with col2:
                # 공정명, 단위작업장소, 제품명 - 자동 매칭
                process_options = ['(없음)'] + list(df.columns)
                unit_options = ['(없음)'] + list(df.columns)
                product_options = ['(없음)'] + list(df.columns)
                content_options = ['(없음)'] + list(df.columns)
                
                # 자동 매칭된 컬럼이 있으면 해당 index, 없으면 0 (없음)
                process_idx = process_options.index(auto_process) if auto_process and auto_process in process_options else 0
                unit_idx = unit_options.index(auto_unit) if auto_unit and auto_unit in unit_options else 0
                product_idx = product_options.index(auto_product) if auto_product and auto_product in product_options else 0
                content_idx = content_options.index(auto_content) if auto_content and auto_content in content_options else 0
                
                process_col = st.selectbox("공정명 컬럼", process_options, index=process_idx)
                unit_col = st.selectbox("단위작업장소 컬럼", unit_options, index=unit_idx)
                product_col = st.selectbox("제품명 컬럼", product_options, index=product_idx)
                content_col = st.selectbox("함유량 컬럼", content_options, index=content_idx)
            
            # 자동 매칭 결과 표시
            if auto_process or auto_unit or auto_product:
                st.success(f"✅ 자동 매칭: 공정명={auto_process or '없음'}, 단위작업장소={auto_unit or '없음'}, 제품명={auto_product or '없음'}, 함유량={auto_content or '없음'}")
            
            # 배치 크기 설정
            batch_size = st.number_input("배치 크기 (한 번에 처리할 행 수)", min_value=50, max_value=500, value=300, step=50)
            
            # 이미 처리된 행 수 추적
            if 'processed_rows' not in st.session_state:
                st.session_state.processed_rows = 0
            
            remaining = len(df) - st.session_state.processed_rows
            st.info(f"📊 총 {len(df)}행 중 **{st.session_state.processed_rows}행 처리 완료**, 남은 행: **{remaining}행**")
            
            col_btn1, col_btn2 = st.columns(2)
            
            with col_btn1:
                if st.button(f"🚀 다음 {min(batch_size, remaining)}건 등록", type="primary", use_container_width=True, disabled=(remaining == 0)):
                    progress = st.progress(0)
                    status = st.empty()
                    
                    success, skip, hazmat_count = 0, 0, 0
                    start_idx = st.session_state.processed_rows
                    end_idx = min(start_idx + batch_size, len(df))
                    batch_total = end_idx - start_idx
                    
                    batch_df = df.iloc[start_idx:end_idx]
                    
                    for i, (idx, row) in enumerate(batch_df.iterrows()):
                        try:
                            # CAS 값 가져오기
                            cas_val = row[cas_col] if cas_col in row.index else None
                            cas = str(cas_val).strip() if cas_val is not None and str(cas_val).strip() not in ['', 'None', 'nan'] else ''
                            
                            if not cas:
                                skip += 1
                                progress.progress((i + 1) / batch_total)
                                continue
                            
                            # 다른 컬럼 값 가져오기
                            def get_val(col_name):
                                if col_name in ['(없음)', '(자동조회)']:
                                    return ''
                                try:
                                    val = row.get(col_name) if hasattr(row, 'get') else row[col_name]
                                    if val is None:
                                        return ''
                                    val_str = str(val).strip()
                                    return '' if val_str in ['', 'None', 'nan'] else val_str
                                except:
                                    return ''
                            
                            chem_name = get_val(name_col) if name_col != '(자동조회)' else ''
                            process = get_val(process_col) if process_col != '(없음)' else ''
                            unit_wp = get_val(unit_col) if unit_col != '(없음)' else ''
                            product = get_val(product_col) if product_col != '(없음)' else ''
                            content = get_val(content_col) if content_col != '(없음)' else ''
                            
                            kosha_data, keco_data, prtr_status = None, None, None
                            
                            if KOSHA_AVAILABLE:
                                status.text(f"[{i+1}/{batch_total}] KOSHA: {cas}")
                                kosha_data, _ = query_chemical_info(cas)
                                try:
                                    prtr_status = check_prtr_status(cas)
                                except:
                                    pass
                            
                            if KECO_AVAILABLE:
                                status.text(f"[{i+1}/{batch_total}] KECO: {cas}")
                                keco_data = get_chemical_regulations(cas)
                            
                            item = create_inventory_item(process, unit_wp, product, chem_name, '', cas, content, kosha_data, keco_data, prtr_status)
                            
                            if item.get('위험물류별', '-') != '-':
                                hazmat_count += 1
                            
                            st.session_state.inventory.append(item)
                            success += 1
                            progress.progress((i + 1) / batch_total)
                        
                        except Exception as e:
                            status.empty()
                            progress.empty()
                            st.error(f"❌ 오류 발생! (행: {start_idx + i + 3}, CAS: {cas_val})")
                            st.code(str(e))
                            st.stop()
                    
                    st.session_state.processed_rows = end_idx
                    status.empty()
                    progress.empty()
                    
                    remaining_after = len(df) - st.session_state.processed_rows
                    if remaining_after > 0:
                        st.success(f"✅ 배치 완료! 성공: {success}건, 건너뜀(CAS없음): {skip}건, 위험물: {hazmat_count}종 | 남은 행: **{remaining_after}건**")
                    else:
                        st.success(f"🎉 전체 완료! 총 등록: **{len(st.session_state.inventory)}종**, 건너뜀: {skip}건, 위험물: {hazmat_count}종")
                        st.session_state.processed_rows = 0
                    st.rerun()
            
            with col_btn2:
                if st.button("🔄 처음부터 다시", use_container_width=True):
                    st.session_state.processed_rows = 0
                    st.session_state.inventory = []
                    st.rerun()
        
        except Exception as e:
            st.error(f"❌ 파일 읽기 오류: {e}")
            import traceback
            st.code(traceback.format_exc())

# ============================================
# 탭 2: 개별 등록
# ============================================
with tab2:
    st.subheader("➕ 개별 등록")
    
    col1, col2 = st.columns(2)
    with col1:
        process = st.text_input("공정명", placeholder="예: 세정공정")
        unit_wp = st.text_input("단위작업장소", placeholder="예: 1라인")
        product = st.text_input("제품명", placeholder="예: 산업용 세정제")
        cas = st.text_input("CAS 번호 *", placeholder="예: 67-64-1")
    with col2:
        content = st.text_input("함유량(%)", placeholder="예: 50")
        alias = st.text_input("관용명", placeholder="예: 아세톤")
        st.markdown("""
        💡 **KOSHA API 자동 조회 항목:**
        - 8번: 노출기준 (TWA, STEL)
        - 15번: 법적규제 + **위험물** 정보
        
        **KECO API (환경부):**
        - 유독물질, 사고대비물질, 제한/금지/허가물질
        """)
    
    if st.button("🔍 조회 및 등록", type="primary", use_container_width=True):
        if cas:
            with st.spinner("API 조회 중..."):
                kosha_data, err = None, None
                keco_data = None
                try:
                    kosha_data, err = query_chemical_info(cas)
                except:
                    pass
                
                prtr_status = check_prtr_status(cas)
                
                if KECO_AVAILABLE:
                    keco_data = get_chemical_regulations(cas)
            
            if kosha_data or (keco_data and keco_data.get('success')):
                chem_name = ''
                if kosha_data:
                    chem_name = kosha_data.get('chemNmKr', '') or kosha_data.get('chemNmEn', '')
                elif keco_data:
                    chem_name = keco_data.get('chemNmKr', '')
                
                item = create_inventory_item(process, unit_wp, product, chem_name, alias, cas, content, kosha_data, keco_data, prtr_status)
                
                if cas.strip() not in [i['CAS No'] for i in st.session_state.inventory]:
                    st.session_state.inventory.append(item)
                    
                    with st.expander("📋 등록된 정보", expanded=True):
                        col_a, col_b = st.columns(2)
                        with col_a:
                            st.write(f"**물질명:** {item['화학물질명']}")
                            st.write(f"**노출기준(TWA):** {item['노출기준(TWA)']}")
                            st.write(f"**작업환경측정:** {item['작업환경측정']}")
                        with col_b:
                            st.write(f"**위험물류별:** {item['위험물류별']}")
                            st.write(f"**지정수량:** {item['지정수량']}")
                            st.write(f"**위험등급:** {item['위험등급']}")
                    
                    st.success(f"✅ **{item['화학물질명']}** 등록 완료!")
                else:
                    st.warning("이미 등록된 CAS 번호입니다.")
            else:
                st.error(f"❌ 조회 실패: {err}")
        else:
            st.warning("CAS 번호를 입력하세요.")

# ============================================
# 탭 3: 목록
# ============================================
with tab3:
    st.subheader("📋 인벤토리 목록")
    
    if st.session_state.inventory:
        df = pd.DataFrame(st.session_state.inventory)
        
        display_cols = ['공정명', '단위작업장소', '제품명', '화학물질명', 'CAS No', '함유량(%)', '노출기준(TWA)', '작업환경측정', '특수건강진단', '위험물류별', '지정수량', '위험등급']
        available_cols = [c for c in display_cols if c in df.columns]
        
        st.dataframe(df[available_cols], use_container_width=True, height=500)
        
        st.divider()
        col1, col2 = st.columns([3, 1])
        with col1:
            del_idx = st.selectbox("삭제할 항목", range(len(st.session_state.inventory)),
                                   format_func=lambda x: f"{st.session_state.inventory[x]['CAS No']} - {st.session_state.inventory[x]['화학물질명']}")
        with col2:
            if st.button("🗑️ 삭제"):
                st.session_state.inventory.pop(del_idx)
                st.rerun()
    else:
        st.info("등록된 물질이 없습니다.")

# ============================================
# 탭 4: 내보내기
# ============================================
with tab4:
    st.subheader("📥 내보내기")
    
    if st.session_state.inventory:
        col1, col2 = st.columns(2)
        with col1:
            excel_file = export_inventory_to_excel(st.session_state.inventory)
            st.download_button("📥 엑셀 다운로드", excel_file, "inventory_result.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col2:
            st.metric("총 물질 수", f"{len(st.session_state.inventory)}종")
        
        st.divider()
        st.subheader("📊 요약")
        
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("작업환경측정", f"{sum(1 for i in st.session_state.inventory if i.get('작업환경측정')=='O')}종")
        c2.metric("특수건강진단", f"{sum(1 for i in st.session_state.inventory if i.get('특수건강진단')=='O')}종")
        c3.metric("관리대상유해물질", f"{sum(1 for i in st.session_state.inventory if i.get('관리대상유해물질')=='O')}종")
        c4.metric("위험물", f"{sum(1 for i in st.session_state.inventory if i.get('위험물류별','-')!='-')}종")
        c5.metric("PRTR대상", f"{sum(1 for i in st.session_state.inventory if i.get('PRTR그룹','-')!='-')}종")
    else:
        st.info("내보낼 데이터가 없습니다.")

st.divider()
st.caption("© 2025 Kay's Chem Manager | KOSHA API + KECO API 연동")
