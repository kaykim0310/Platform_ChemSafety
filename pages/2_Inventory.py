#!/usr/bin/env python3
"""
📦 인벤토리 관리 시스템
- KOSHA API 연동 (서버 복구 시 자동 작동)
- 템플릿 서식 적용 (2행 헤더)
- 규제정보 자동 조회
"""
import streamlit as st
import pandas as pd
import sys
from pathlib import Path
from datetime import date, datetime
import io

# 경로 설정
current_dir = Path(__file__).parent.parent
sys.path.insert(0, str(current_dir))

# KOSHA API 모듈 import
try:
    from core.kosha_api import get_full_msds_data, search_by_cas
    from core.prtr_db import check_prtr_status
    KOSHA_AVAILABLE = True
except ImportError:
    KOSHA_AVAILABLE = False

# ============================================
# 페이지 설정
# ============================================
st.set_page_config(
    page_title="인벤토리 관리",
    page_icon="📦",
    layout="wide"
)

# ============================================
# 스타일
# ============================================
st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1.5rem;
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        border-radius: 10px;
        margin-bottom: 1.5rem;
        color: white;
    }
    .api-status-ok {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        background: #dcfce7;
        color: #166534;
        border-radius: 1rem;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .api-status-error {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        background: #fee2e2;
        color: #991b1b;
        border-radius: 1rem;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .reg-yes { color: #dc2626; font-weight: bold; }
    .reg-no { color: #6b7280; }
</style>
""", unsafe_allow_html=True)

# ============================================
# 세션 상태 초기화
# ============================================
if 'inventory' not in st.session_state:
    st.session_state.inventory = []

if 'api_test_result' not in st.session_state:
    st.session_state.api_test_result = None

# ============================================
# 유틸리티 함수
# ============================================
def test_kosha_api():
    """KOSHA API 연결 테스트"""
    if not KOSHA_AVAILABLE:
        return False, "모듈 로드 실패"
    
    try:
        result = search_by_cas("67-64-1")  # 아세톤으로 테스트
        if result and result.get('chemId'):
            return True, "연결 성공"
        else:
            return False, "응답 없음 (서버 점검 중)"
    except Exception as e:
        return False, f"연결 실패: {str(e)[:30]}"

def get_chemical_info(cas_no):
    """CAS 번호로 화학물질 정보 조회"""
    if not KOSHA_AVAILABLE:
        return None, "KOSHA 모듈 없음"
    
    try:
        result = get_full_msds_data(cas_no)
        if result.get('success'):
            return result, None
        else:
            return None, result.get('error', '조회 실패')
    except Exception as e:
        return None, f"API 오류: {str(e)[:50]}"

def extract_carcinogenicity(kosha_data):
    """발암성 정보 추출"""
    if not kosha_data:
        return "-"
    
    tox = kosha_data.get('toxicity_info', {})
    iarc = tox.get('IARC', '-')
    
    if 'Group 1' in str(iarc):
        return "1군(확인)"
    elif 'Group 2A' in str(iarc):
        return "2A군(추정)"
    elif 'Group 2B' in str(iarc):
        return "2B군(가능)"
    elif iarc and iarc != '-':
        return iarc[:10]
    
    # GHS 분류에서 확인
    hazard = kosha_data.get('hazard_classification', {})
    ghs = hazard.get('ghs_classification', [])
    for g in ghs:
        if '발암성' in str(g):
            if '구분1' in str(g) or '1A' in str(g) or '1B' in str(g):
                return "1군"
            elif '구분2' in str(g):
                return "2군"
    
    return "-"

def extract_mutagenicity(kosha_data):
    """변이원성 정보 추출"""
    if not kosha_data:
        return "-"
    
    hazard = kosha_data.get('hazard_classification', {})
    ghs = hazard.get('ghs_classification', [])
    for g in ghs:
        if '변이원성' in str(g) or '생식세포 변이' in str(g):
            if '구분1' in str(g):
                return "O"
            elif '구분2' in str(g):
                return "△"
    return "-"

def extract_reproductive_toxicity(kosha_data):
    """생식독성 정보 추출"""
    if not kosha_data:
        return "-"
    
    hazard = kosha_data.get('hazard_classification', {})
    ghs = hazard.get('ghs_classification', [])
    for g in ghs:
        if '생식독성' in str(g):
            if '구분1' in str(g):
                return "O"
            elif '구분2' in str(g):
                return "△"
    return "-"

def create_inventory_item(process_name, product_name, chem_name, alias, cas_no, content, kosha_data, prtr_status):
    """인벤토리 항목 생성"""
    
    # 기본값
    item = {
        '공정명': process_name,
        '제품명': product_name,
        '화학물질명': chem_name,
        '관용명/이명': alias,
        'CAS No': cas_no,
        '함유량(%)': content,
        # 독성정보
        '발암성': '-',
        '변이성': '-',
        '생식독성': '-',
        '노출기준(TWA)': '-',
        # 법적규제 (산안법)
        '작업환경측정': 'X',
        '특수건강진단': 'X',
        '관리대상유해물질': 'X',
        '특별관리물질': 'X',
        # 위험물
        '위험물류별': '-',
        '지정수량': '-',
        '위험등급': '-',
        # 환경부
        '유독': 'X',
        '사고대비': 'X',
        '제한/금지/허가': '-',
        # PRTR
        'PRTR그룹': '-',
        'PRTR기준량': '-'
    }
    
    # KOSHA 데이터가 있으면 채우기
    if kosha_data:
        # 물질명
        item['화학물질명'] = kosha_data.get('name_kor', chem_name) or chem_name
        
        # 독성정보
        item['발암성'] = extract_carcinogenicity(kosha_data)
        item['변이성'] = extract_mutagenicity(kosha_data)
        item['생식독성'] = extract_reproductive_toxicity(kosha_data)
        
        # 노출기준
        exp = kosha_data.get('exposure_limits', {})
        item['노출기준(TWA)'] = exp.get('TWA', '-')
        
        # 법적규제 (산안법)
        regs = kosha_data.get('legal_regulations', {})
        item['작업환경측정'] = regs.get('작업환경측정', 'X')
        item['특수건강진단'] = regs.get('특수건강진단', 'X')
        item['관리대상유해물질'] = regs.get('관리대상유해물질', 'X')
        item['특별관리물질'] = regs.get('특별관리물질', 'X')
        
        # 위험물 - KOSHA 데이터에서 추출 시도
        # (실제로는 별도 DB 필요, 여기서는 기본값)
        
        # 환경부
        if regs.get('유독물질') and regs.get('유독물질') != '-':
            item['유독'] = 'O'
        if regs.get('사고대비물질') and regs.get('사고대비물질') != '-':
            item['사고대비'] = 'O'
    
    # PRTR 상태
    if prtr_status and prtr_status.get('대상여부') == 'O':
        item['PRTR그룹'] = prtr_status.get('그룹', '-')
        item['PRTR기준량'] = prtr_status.get('기준취급량', '-')
    
    return item

def create_template_excel():
    """템플릿 엑셀 파일 생성 (2행 헤더)"""
    output = io.BytesIO()
    
    # openpyxl로 직접 생성 (병합 셀 포함)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "화학물질 정보"
    
    # 스타일 정의
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    header_fill2 = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    
    # 1행 헤더 (대분류)
    headers_row1 = [
        ('A', 'A', '공정명'),
        ('B', 'B', '제품명'),
        ('C', 'C', '화학물질명'),
        ('D', 'D', '관용명/이명'),
        ('E', 'E', 'CAS No'),
        ('F', 'F', '함유량(%)'),
        ('G', 'J', '독성정보'),
        ('K', 'N', '법적규제 대상여부'),
        ('O', 'Q', '위험물'),
        ('R', 'Z', '환경부 법적규제 대상여부'),
    ]
    
    # 2행 헤더 (세부항목)
    headers_row2 = [
        'A', 'B', 'C', 'D', 'E', 'F',  # 기본 (1행과 병합)
        '발암성', '변이성', '생식독성', '노출기준(TWA)',  # 독성정보 G-J
        '작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질',  # 법적규제 K-N
        '위험물류별', '지정수량', '위험등급',  # 위험물 O-Q
        '기존', '유독', '사고대비', '제한/금지/허가', '중점', '잔류', '함량 및 규제정보', '등록대상기존화학물질', '기존물질여부'  # 환경부 R-Z
    ]
    
    # 1행 작성
    ws['A1'] = '공정명'
    ws['B1'] = '제품명'
    ws['C1'] = '화학물질명'
    ws['D1'] = '관용명/이명'
    ws['E1'] = 'CAS No'
    ws['F1'] = '함유량(%)'
    ws['G1'] = '독성정보'
    ws['K1'] = '법적규제 대상여부'
    ws['O1'] = '위험물'
    ws['R1'] = '환경부 법적규제 대상여부'
    
    # 2행 작성
    row2_headers = [
        '', '', '', '', '', '',  # A-F (병합됨)
        '발암성', '변이성', '생식독성', '노출기준(TWA)',  # G-J
        '작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질',  # K-N
        '위험물류별', '지정수량', '위험등급',  # O-Q
        '기존', '유독', '사고대비', '제한/금지/허가', '중점', '잔류', '함량 및 규제정보', '등록대상기존화학물질', '기존물질여부'  # R-Z
    ]
    
    for col, header in enumerate(row2_headers, 1):
        ws.cell(row=2, column=col, value=header)
    
    # 셀 병합
    # A-F: 1행~2행 병합
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.merge_cells(f'{col}1:{col}2')
    
    # 대분류 병합
    ws.merge_cells('G1:J1')  # 독성정보
    ws.merge_cells('K1:N1')  # 법적규제 대상여부
    ws.merge_cells('O1:Q1')  # 위험물
    ws.merge_cells('R1:Z1')  # 환경부 법적규제 대상여부
    
    # 스타일 적용
    for row in [1, 2]:
        for col in range(1, 27):  # A~Z
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill if row == 1 else header_fill2
    
    # 열 너비 설정
    col_widths = {
        'A': 12, 'B': 15, 'C': 20, 'D': 15, 'E': 12, 'F': 10,
        'G': 8, 'H': 8, 'I': 8, 'J': 15,
        'K': 12, 'L': 12, 'M': 14, 'N': 12,
        'O': 12, 'P': 10, 'Q': 10,
        'R': 8, 'S': 8, 'T': 8, 'U': 12, 'V': 8, 'W': 8, 'X': 14, 'Y': 16, 'Z': 12
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 행 높이
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    
    wb.save(output)
    output.seek(0)
    return output

def export_inventory_to_excel(inventory_data):
    """인벤토리 데이터를 템플릿 형식으로 내보내기"""
    output = io.BytesIO()
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "화학물질 정보"
    
    # 스타일
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    header_fill2 = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    yes_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # 1행 헤더
    ws['A1'] = '공정명'
    ws['B1'] = '제품명'
    ws['C1'] = '화학물질명'
    ws['D1'] = '관용명/이명'
    ws['E1'] = 'CAS No'
    ws['F1'] = '함유량(%)'
    ws['G1'] = '독성정보'
    ws['K1'] = '법적규제 대상여부'
    ws['O1'] = '위험물'
    ws['R1'] = '환경부 법적규제 대상여부'
    
    # 2행 헤더
    row2_headers = [
        '', '', '', '', '', '',
        '발암성', '변이성', '생식독성', '노출기준(TWA)',
        '작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질',
        '위험물류별', '지정수량', '위험등급',
        '기존', '유독', '사고대비', '제한/금지/허가', '중점', '잔류', '함량 및 규제정보', '등록대상기존화학물질', '기존물질여부'
    ]
    
    for col, header in enumerate(row2_headers, 1):
        ws.cell(row=2, column=col, value=header)
    
    # 셀 병합
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.merge_cells(f'{col}1:{col}2')
    ws.merge_cells('G1:J1')
    ws.merge_cells('K1:N1')
    ws.merge_cells('O1:Q1')
    ws.merge_cells('R1:Z1')
    
    # 헤더 스타일
    for row in [1, 2]:
        for col in range(1, 27):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill if row == 1 else header_fill2
    
    # 데이터 입력
    for row_idx, item in enumerate(inventory_data, 3):
        data_row = [
            item.get('공정명', ''),
            item.get('제품명', ''),
            item.get('화학물질명', ''),
            item.get('관용명/이명', ''),
            item.get('CAS No', ''),
            item.get('함유량(%)', ''),
            item.get('발암성', '-'),
            item.get('변이성', '-'),
            item.get('생식독성', '-'),
            item.get('노출기준(TWA)', '-'),
            item.get('작업환경측정', 'X'),
            item.get('특수건강진단', 'X'),
            item.get('관리대상유해물질', 'X'),
            item.get('특별관리물질', 'X'),
            item.get('위험물류별', '-'),
            item.get('지정수량', '-'),
            item.get('위험등급', '-'),
            '-',  # 기존
            item.get('유독', 'X'),
            item.get('사고대비', 'X'),
            item.get('제한/금지/허가', '-'),
            '-',  # 중점
            '-',  # 잔류
            '-',  # 함량 및 규제정보
            '-',  # 등록대상기존화학물질
            '-',  # 기존물질여부
        ]
        
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
            
            # 'O' 값에 빨간 배경
            if value == 'O':
                cell.fill = yes_fill
    
    # 열 너비
    col_widths = {
        'A': 12, 'B': 15, 'C': 20, 'D': 15, 'E': 12, 'F': 10,
        'G': 8, 'H': 8, 'I': 8, 'J': 15,
        'K': 12, 'L': 12, 'M': 14, 'N': 12,
        'O': 12, 'P': 10, 'Q': 10,
        'R': 8, 'S': 8, 'T': 8, 'U': 12, 'V': 8, 'W': 8, 'X': 14, 'Y': 16, 'Z': 12
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output)
    output.seek(0)
    return output

# ============================================
# 사이드바
# ============================================
with st.sidebar:
    st.markdown("### 📦 인벤토리 관리")
    
    # API 상태 확인
    if st.button("🔌 API 연결 테스트", use_container_width=True):
        with st.spinner("테스트 중..."):
            success, msg = test_kosha_api()
            st.session_state.api_test_result = (success, msg)
    
    if st.session_state.api_test_result:
        success, msg = st.session_state.api_test_result
        if success:
            st.markdown(f'<span class="api-status-ok">✅ {msg}</span>', unsafe_allow_html=True)
        else:
            st.markdown(f'<span class="api-status-error">❌ {msg}</span>', unsafe_allow_html=True)
    
    st.divider()
    
    # 통계
    inv_count = len(st.session_state.inventory)
    st.metric("등록된 물질", f"{inv_count}종")
    
    if inv_count > 0:
        measurement_count = sum(1 for i in st.session_state.inventory if i.get('작업환경측정') == 'O')
        st.metric("측정대상 물질", f"{measurement_count}종")
    
    st.divider()
    
    # 템플릿 다운로드
    st.markdown("#### 📥 템플릿")
    template_data = create_template_excel()
    st.download_button(
        "📄 빈 템플릿 다운로드",
        data=template_data.getvalue(),
        file_name=f"인벤토리_템플릿_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.divider()
    
    # 초기화
    if st.button("🗑️ 전체 삭제", use_container_width=True):
        st.session_state.inventory = []
        st.rerun()

# ============================================
# 메인 컨텐츠
# ============================================
st.markdown("""
<div class="main-header">
    <h2>📦 화학물질 인벤토리 관리</h2>
    <p>CAS 번호 입력 → KOSHA API 자동 조회 → 규제정보 확인</p>
</div>
""", unsafe_allow_html=True)

# 탭 구성
tab1, tab2, tab3 = st.tabs(["➕ 물질 등록", "📋 인벤토리 목록", "📤 내보내기"])

# ============================================
# 탭 1: 물질 등록
# ============================================
with tab1:
    st.subheader("화학물질 등록")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 기본 정보")
        process_name = st.text_input("공정명", placeholder="예: 세정공정")
        product_name = st.text_input("제품명", placeholder="예: 산업용 세정제")
        cas_input = st.text_input("CAS 번호 *", placeholder="예: 67-64-1")
        content = st.text_input("함유량(%)", placeholder="예: 50 또는 40~60")
    
    with col2:
        st.markdown("#### 추가 정보 (선택)")
        alias = st.text_input("관용명/이명", placeholder="예: 아세톤, 디메틸케톤")
        
        st.info("""
        💡 **CAS 번호만 입력하면 자동 조회됩니다!**
        - 화학물질명, 노출기준
        - 발암성, 변이성, 생식독성
        - 작업환경측정/특수건강진단 대상
        - 관리대상유해물질/특별관리물질
        """)
    
    st.divider()
    
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        if st.button("🔍 KOSHA 조회 및 등록", type="primary", use_container_width=True):
            if cas_input:
                with st.spinner(f"'{cas_input}' 조회 중..."):
                    kosha_data, error = get_chemical_info(cas_input.strip())
                    prtr_status = check_prtr_status(cas_input.strip()) if KOSHA_AVAILABLE else None
                
                if kosha_data:
                    chem_name = kosha_data.get('name_kor', '')
                    
                    # 인벤토리 항목 생성
                    item = create_inventory_item(
                        process_name=process_name,
                        product_name=product_name,
                        chem_name=chem_name,
                        alias=alias,
                        cas_no=cas_input.strip(),
                        content=content,
                        kosha_data=kosha_data,
                        prtr_status=prtr_status
                    )
                    
                    # 중복 체크
                    existing = [i['CAS No'] for i in st.session_state.inventory]
                    if cas_input.strip() in existing:
                        st.warning("⚠️ 이미 등록된 물질입니다.")
                    else:
                        st.session_state.inventory.append(item)
                        st.success(f"✅ **{chem_name}** 등록 완료!")
                        
                        # 조회 결과 요약
                        st.markdown("#### 📊 조회 결과")
                        col_a, col_b, col_c, col_d = st.columns(4)
                        with col_a:
                            st.metric("노출기준(TWA)", item['노출기준(TWA)'])
                        with col_b:
                            val = "대상" if item['작업환경측정'] == 'O' else "비대상"
                            st.metric("작업환경측정", val)
                        with col_c:
                            val = "대상" if item['특수건강진단'] == 'O' else "비대상"
                            st.metric("특수건강진단", val)
                        with col_d:
                            st.metric("발암성", item['발암성'])
                        
                        st.rerun()
                else:
                    st.error(f"❌ 조회 실패: {error}")
                    st.info("💡 KOSHA API 서버가 복구 중일 수 있습니다. 나중에 다시 시도해주세요.")
            else:
                st.warning("CAS 번호를 입력하세요.")
    
    # 테스트용 CAS 번호
    st.divider()
    with st.expander("📚 테스트용 CAS 번호 예시"):
        st.markdown("""
        | CAS 번호 | 물질명 | 특징 |
        |----------|--------|------|
        | `67-64-1` | 아세톤 | 제1석유류, 작업환경측정 대상 |
        | `108-88-3` | 톨루엔 | 유기용제, 작업환경측정+특수건강진단 |
        | `71-43-2` | 벤젠 | 1군 발암물질, 특별관리물질 |
        | `50-00-0` | 포름알데히드 | 1군 발암물질 |
        | `110-54-3` | n-헥산 | 말초신경 독성 |
        """)

# ============================================
# 탭 2: 인벤토리 목록
# ============================================
with tab2:
    st.subheader("등록된 화학물질 목록")
    
    if st.session_state.inventory:
        # 필터
        col1, col2, col3 = st.columns(3)
        with col1:
            filter_measurement = st.checkbox("작업환경측정 대상만")
        with col2:
            filter_health = st.checkbox("특수건강진단 대상만")
        with col3:
            filter_cmr = st.checkbox("CMR 물질만")
        
        # 필터 적용
        filtered = st.session_state.inventory.copy()
        if filter_measurement:
            filtered = [i for i in filtered if i.get('작업환경측정') == 'O']
        if filter_health:
            filtered = [i for i in filtered if i.get('특수건강진단') == 'O']
        if filter_cmr:
            filtered = [i for i in filtered if i.get('발암성') not in ['-', ''] or i.get('변이성') == 'O' or i.get('생식독성') == 'O']
        
        st.caption(f"총 {len(filtered)}종 표시")
        
        # 테이블 표시
        if filtered:
            # 주요 컬럼만 표시
            display_cols = ['CAS No', '화학물질명', '함유량(%)', '노출기준(TWA)', 
                          '작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질',
                          '발암성', '변이성', '생식독성']
            
            df = pd.DataFrame(filtered)[display_cols]
            
            # 스타일 적용
            def highlight_yes(val):
                if val == 'O':
                    return 'background-color: #fee2e2; color: #991b1b; font-weight: bold'
                return ''
            
            styled_df = df.style.applymap(highlight_yes, subset=['작업환경측정', '특수건강진단', '관리대상유해물질', '특별관리물질'])
            
            st.dataframe(styled_df, use_container_width=True, height=400)
            
            # 삭제 기능
            st.divider()
            col1, col2 = st.columns([3, 1])
            with col1:
                delete_idx = st.selectbox(
                    "삭제할 물질 선택",
                    range(len(st.session_state.inventory)),
                    format_func=lambda x: f"{st.session_state.inventory[x]['CAS No']} - {st.session_state.inventory[x]['화학물질명']}"
                )
            with col2:
                if st.button("🗑️ 삭제", use_container_width=True):
                    st.session_state.inventory.pop(delete_idx)
                    st.rerun()
        else:
            st.info("필터 조건에 맞는 물질이 없습니다.")
    else:
        st.info("등록된 화학물질이 없습니다. '물질 등록' 탭에서 추가하세요.")

# ============================================
# 탭 3: 내보내기
# ============================================
with tab3:
    st.subheader("인벤토리 내보내기")
    
    if st.session_state.inventory:
        st.markdown(f"**등록된 물질: {len(st.session_state.inventory)}종**")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📊 엑셀 (템플릿 형식)")
            excel_data = export_inventory_to_excel(st.session_state.inventory)
            st.download_button(
                "📥 엑셀 다운로드",
                data=excel_data.getvalue(),
                file_name=f"인벤토리_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.caption("템플릿 서식 (2행 헤더, 병합 셀 포함)")
        
        with col2:
            st.markdown("#### 📄 CSV (간단 형식)")
            df = pd.DataFrame(st.session_state.inventory)
            csv_data = df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                "📥 CSV 다운로드",
                data=csv_data,
                file_name=f"인벤토리_{date.today()}.csv",
                mime="text/csv",
                use_container_width=True
            )
            st.caption("단순 테이블 형식")
        
        # 통계 요약
        st.divider()
        st.markdown("#### 📈 규제 현황 요약")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            cnt = sum(1 for i in st.session_state.inventory if i.get('작업환경측정') == 'O')
            st.metric("작업환경측정 대상", f"{cnt}종")
        with col2:
            cnt = sum(1 for i in st.session_state.inventory if i.get('특수건강진단') == 'O')
            st.metric("특수건강진단 대상", f"{cnt}종")
        with col3:
            cnt = sum(1 for i in st.session_state.inventory if i.get('관리대상유해물질') == 'O')
            st.metric("관리대상유해물질", f"{cnt}종")
        with col4:
            cnt = sum(1 for i in st.session_state.inventory if i.get('특별관리물질') == 'O')
            st.metric("특별관리물질", f"{cnt}종")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            cnt = sum(1 for i in st.session_state.inventory if i.get('발암성') not in ['-', ''])
            st.metric("발암성 물질", f"{cnt}종")
        with col2:
            cnt = sum(1 for i in st.session_state.inventory if i.get('유독') == 'O')
            st.metric("유독물질", f"{cnt}종")
        with col3:
            cnt = sum(1 for i in st.session_state.inventory if i.get('PRTR그룹') != '-')
            st.metric("PRTR 대상", f"{cnt}종")
        with col4:
            st.metric("총 등록 물질", f"{len(st.session_state.inventory)}종")
    else:
        st.info("내보낼 데이터가 없습니다. 먼저 화학물질을 등록하세요.")

# ============================================
# 푸터
# ============================================
st.divider()
st.caption("© 2025 화학물질 인벤토리 관리 시스템 | Kay's Chem Manager | KOSHA API 연동")
